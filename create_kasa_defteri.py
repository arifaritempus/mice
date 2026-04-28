#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Restoran Kasa Defteri Excel Oluşturucu
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

def create_restaurant_cash_register():
    """Restoran kasa defteri Excel dosyası oluşturur"""
    
    wb = Workbook()
    
    # 1. GÜNLÜK İŞLEMLER sayfası
    ws_daily = wb.active
    ws_daily.title = "Günlük İşlemler"
    
    # Başlıklar
    headers = [
        "Tarih", "Fiş No", "Açıklama", 
        "Nakit Giriş", "Nakit Çıkış",
        "POS Giriş", "POS Çıkış",
        "Havale Giriş", "Havale Çıkış",
        "Çek/Senet Giriş", "Çek/Senet Çıkış",
        "Kredili Satış", "Kredili Tahsilat",
        "Müşteri Adı", "Toplam Giriş", "Toplam Çıkış", "Kasa Bakiye", "Notlar"
    ]
    
    # Başlık stilini ayarla
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col_num, header in enumerate(headers, 1):
        cell = ws_daily.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # Örnek veriler
    sample_data = [
        ["01.01.2026", "001", "Açılış Bakiyesi", 5000, 0, 0, 0, 0, 0, 0, 0, 0, 0, "-", "=D2+F2+H2+J2+L2+M2", "=E2+G2+I2+K2", "=Q1+O2-P2", "Kasa açılış"],
        ["01.01.2026", "002", "Masa 1 - Nakit", 450, 0, 0, 0, 0, 0, 0, 0, 0, 0, "-", "=D3+F3+H3+J3+L3+M3", "=E3+G3+I3+K3", "=Q2+O3-P3", ""],
        ["01.01.2026", "003", "Masa 2 - POS", 0, 0, 680, 0, 0, 0, 0, 0, 0, 0, "-", "=D4+F4+H4+J4+L4+M4", "=E4+G4+I4+K4", "=Q3+O4-P4", ""],
        ["01.01.2026", "004", "Masa 3 - Kredili", 0, 0, 0, 0, 0, 0, 0, 0, 1250, 0, "Ahmet Yılmaz", "=D5+F5+H5+J5+L5+M5", "=E5+G5+I5+K5", "=Q4+O5-P5", "Veresiye"],
        ["01.01.2026", "005", "Sebze Alımı", 0, 850, 0, 0, 0, 0, 0, 0, 0, 0, "-", "=D6+F6+H6+J6+L6+M6", "=E6+G6+I6+K6", "=Q5+O6-P6", "Pazar"],
        ["01.01.2026", "006", "Masa 4 - Havale", 0, 0, 0, 0, 920, 0, 0, 0, 0, 0, "-", "=D7+F7+H7+J7+L7+M7", "=E7+G7+I7+K7", "=Q6+O7-P7", ""],
        ["01.01.2026", "007", "Elektrik Faturası", 0, 0, 0, 1200, 0, 0, 0, 0, 0, 0, "-", "=D8+F8+H8+J8+L8+M8", "=E8+G8+I8+K8", "=Q7+O8-P8", ""],
        ["01.01.2026", "008", "Masa 5 - Nakit", 320, 0, 0, 0, 0, 0, 0, 0, 0, 0, "-", "=D9+F9+H9+J9+L9+M9", "=E9+G9+I9+K9", "=Q8+O9-P9", ""],
        ["01.01.2026", "009", "Kredili Tahsilat", 0, 0, 0, 0, 0, 0, 0, 0, 0, 500, "Mehmet Demir", "=D10+F10+H10+J10+L10+M10", "=E10+G10+I10+K10", "=Q9+O10-P10", "Kısmi ödeme"],
        ["01.01.2026", "010", "Personel Maaş", 0, 3000, 0, 0, 0, 0, 0, 0, 0, 0, "-", "=D11+F11+H11+J11+L11+M11", "=E11+G11+I11+K11", "=Q10+O11-P11", "Garson"],
    ]
    
    # Verileri ekle
    for row_num, row_data in enumerate(sample_data, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws_daily.cell(row=row_num, column=col_num)
            if isinstance(value, str) and value.startswith("="):
                cell.value = value
            else:
                cell.value = value
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center' if col_num in [1, 2, 14] else 'left', vertical='center')
            
            # Para formatı
            if col_num in [4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 15, 16, 17]:
                cell.number_format = '#,##0.00 ₺'
    
    # Günlük özet satırı
    summary_row = 13
    ws_daily.cell(row=summary_row, column=1).value = "GÜNLÜK ÖZET"
    ws_daily.cell(row=summary_row, column=1).font = Font(bold=True, size=12)
    ws_daily.cell(row=summary_row, column=1).fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    
    summary_formulas = [
        ("", 1), ("", 2), ("TOPLAM:", 3),
        ("=SUM(D2:D11)", 4), ("=SUM(E2:E11)", 5),
        ("=SUM(F2:F11)", 6), ("=SUM(G2:G11)", 7),
        ("=SUM(H2:H11)", 8), ("=SUM(I2:I11)", 9),
        ("=SUM(J2:J11)", 10), ("=SUM(K2:K11)", 11),
        ("=SUM(L2:L11)", 12), ("=SUM(M2:M11)", 13),
        ("", 14), ("=SUM(O2:O11)", 15), ("=SUM(P2:P11)", 16),
        ("=Q11", 17), ("", 18)
    ]
    
    for formula, col_num in summary_formulas:
        cell = ws_daily.cell(row=summary_row, column=col_num)
        cell.value = formula
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        cell.border = thin_border
        if col_num in [4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 15, 16, 17]:
            cell.number_format = '#,##0.00 ₺'
    
    # Sütun genişliklerini ayarla
    column_widths = [12, 8, 20, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 15, 12, 12, 12, 15]
    for i, width in enumerate(column_widths, 1):
        ws_daily.column_dimensions[get_column_letter(i)].width = width
    
    # 2. KREDİLİ HESAPLAR sayfası
    ws_credit = wb.create_sheet("Kredili Hesaplar")
    
    credit_headers = ["Müşteri Adı", "Toplam Borç", "Ödenen", "Kalan Borç", "Son İşlem Tarihi", "Telefon", "Notlar"]
    
    for col_num, header in enumerate(credit_headers, 1):
        cell = ws_credit.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    credit_data = [
        ["Ahmet Yılmaz", 1250, 0, "=B2-C2", "01.01.2026", "0532 123 4567", ""],
        ["Mehmet Demir", 2000, 500, "=B3-C3", "01.01.2026", "0533 234 5678", ""],
        ["Ayşe Kaya", 750, 0, "=B4-C4", "28.12.2025", "0534 345 6789", ""],
        ["Fatma Öz", 1500, 1000, "=B5-C5", "30.12.2025", "0535 456 7890", ""],
    ]
    
    for row_num, row_data in enumerate(credit_data, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws_credit.cell(row=row_num, column=col_num)
            if isinstance(value, str) and value.startswith("="):
                cell.value = value
            else:
                cell.value = value
            cell.border = thin_border
            if col_num in [2, 3, 4]:
                cell.number_format = '#,##0.00 ₺'
    
    # Toplam satırı
    total_row = len(credit_data) + 3
    ws_credit.cell(row=total_row, column=1).value = "TOPLAM"
    ws_credit.cell(row=total_row, column=1).font = Font(bold=True)
    ws_credit.cell(row=total_row, column=1).fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    
    for col_num in [2, 3, 4]:
        cell = ws_credit.cell(row=total_row, column=col_num)
        cell.value = f"=SUM({get_column_letter(col_num)}2:{get_column_letter(col_num)}{total_row-1})"
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        cell.number_format = '#,##0.00 ₺'
        cell.border = thin_border
    
    credit_widths = [20, 15, 15, 15, 15, 15, 25]
    for i, width in enumerate(credit_widths, 1):
        ws_credit.column_dimensions[get_column_letter(i)].width = width
    
    # 3. AYLIK RAPOR sayfası
    ws_monthly = wb.create_sheet("Aylık Rapor")
    
    ws_monthly.cell(row=1, column=1).value = "AYLIK KASA RAPORU"
    ws_monthly.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws_monthly.merge_cells('A1:E1')
    ws_monthly.cell(row=1, column=1).alignment = Alignment(horizontal='center')
    
    monthly_headers = ["Ödeme Yöntemi", "Giriş", "Çıkış", "Net", "Yüzde"]
    for col_num, header in enumerate(monthly_headers, 1):
        cell = ws_monthly.cell(row=3, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    monthly_data = [
        ["Nakit", "=SUM('Günlük İşlemler'!D:D)", "=SUM('Günlük İşlemler'!E:E)", "=B4-C4", "=D4/$D$10"],
        ["POS", "=SUM('Günlük İşlemler'!F:F)", "=SUM('Günlük İşlemler'!G:G)", "=B5-C5", "=D5/$D$10"],
        ["Havale", "=SUM('Günlük İşlemler'!H:H)", "=SUM('Günlük İşlemler'!I:I)", "=B6-C6", "=D6/$D$10"],
        ["Çek/Senet", "=SUM('Günlük İşlemler'!J:J)", "=SUM('Günlük İşlemler'!K:K)", "=B7-C7", "=D7/$D$10"],
        ["Kredili Satış", "=SUM('Günlük İşlemler'!L:L)", "0", "=B8-C8", "=D8/$D$10"],
        ["Kredili Tahsilat", "=SUM('Günlük İşlemler'!M:M)", "0", "=B9-C9", "=D9/$D$10"],
    ]
    
    for row_num, row_data in enumerate(monthly_data, 4):
        for col_num, value in enumerate(row_data, 1):
            cell = ws_monthly.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = thin_border
            if col_num in [2, 3, 4]:
                cell.number_format = '#,##0.00 ₺'
            elif col_num == 5:
                cell.number_format = '0.00%'
    
    # Toplam satırı
    total_row = 10
    ws_monthly.cell(row=total_row, column=1).value = "TOPLAM"
    ws_monthly.cell(row=total_row, column=1).font = Font(bold=True)
    ws_monthly.cell(row=total_row, column=1).fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    
    for col_num in [2, 3, 4]:
        cell = ws_monthly.cell(row=total_row, column=col_num)
        cell.value = f"=SUM({get_column_letter(col_num)}4:{get_column_letter(col_num)}9)"
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        cell.number_format = '#,##0.00 ₺'
        cell.border = thin_border
    
    cell = ws_monthly.cell(row=total_row, column=5)
    cell.value = "100%"
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    cell.border = thin_border
    
    monthly_widths = [20, 15, 15, 15, 12]
    for i, width in enumerate(monthly_widths, 1):
        ws_monthly.column_dimensions[get_column_letter(i)].width = width
    
    # Dosyayı kaydet
    output_file = "/Users/arifari/Desktop/TT_Sistem_AG/Restoran_Kasa_Defteri.xlsx"
    wb.save(output_file)
    print(f"✅ Excel dosyası başarıyla oluşturuldu: {output_file}")
    return output_file

if __name__ == "__main__":
    create_restaurant_cash_register()
